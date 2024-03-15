import os
from openpyxl import Workbook
from openpyxl.styles import Font

# Define el directorio de inicio como el directorio actual donde se ejecuta el script
dir_path = os.getcwd()
excel_filename = os.path.join(dir_path, "lista_de_archivos.xlsx")

# Crea un nuevo libro de Excel y selecciona la hoja activa
wb = Workbook()
ws = wb.active

# Añade los encabezados a las columnas
ws.append(["Ruta del Archivo", "Nombre del Archivo", "Vínculo al Archivo"])

# Recorre el directorio de forma recursiva
for root, dirs, files in os.walk(dir_path):
    for file in files:
        # Construye la ruta completa del archivo
        file_path = os.path.join(root, file)
        # Construye el vínculo al archivo (ajusta esto según sea necesario)
        file_link = f"file:///{file_path.replace(os.path.sep, '/')}"
        # Añade la ruta del archivo, el nombre del archivo y el vínculo al Excel
        ws.append([file_path, file, file_link])

        # Obtiene la fila actual (la última fila con datos)
        current_row = ws.max_row
        # Establece el enlace como un hipervínculo en la celda
        ws[f'C{current_row}'].hyperlink = file_link
        # Opcional: cambia el estilo de la fuente para que parezca un enlace
        ws[f'C{current_row}'].font = Font(color="0000FF", underline="single")

# Guarda el libro de Excel
wb.save(excel_filename)

print(f"El archivo de Excel ha sido guardado en: {excel_filename}")
