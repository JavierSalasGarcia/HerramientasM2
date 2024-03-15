import openpyxl
import re

# Función para leer el cuestionario y obtener las preguntas y respuestas correctas
def leer_cuestionario(ruta_cuestionario):
    preguntas_y_respuestas = []
    with open(ruta_cuestionario, "r", encoding="utf-8") as archivo:
        lineas = archivo.readlines()[2:]  # Saltar título e indicaciones
        pregunta_actual = []
        for linea in lineas:
            if linea.strip() == "":
                if pregunta_actual:
                    preguntas_y_respuestas.append(pregunta_actual)
                    pregunta_actual = []
            else:
                pregunta_actual.append(linea.strip())
        if pregunta_actual:  # Añadir la última pregunta si existe
            preguntas_y_respuestas.append(pregunta_actual)
    return preguntas_y_respuestas

# Pedir al usuario confirmar la columna inicial o especificar una nueva
columna_predeterminada = "L"  # Columna inicial predeterminada
confirmacion_usuario = input(f"Por defecto, las respuestas comienzan en la columna {columna_predeterminada}. ¿Es correcto? (s/n): ").strip().lower()

if confirmacion_usuario == "s":
    columna_inicial = columna_predeterminada
else:
    columna_inicial = input("Por favor, introduce la columna en la que comienzan las respuestas: ").strip().upper()

# Convertir la columna a un índice numérico (considerando que A=1)
indice_columna_inicial = openpyxl.utils.column_index_from_string(columna_inicial)

# Leer las respuestas del Excel
ruta_excel = "C:/M2S2/Respuestas.xlsx"
wb = openpyxl.load_workbook(ruta_excel)
ws = wb.active

# Leer el cuestionario
ruta_cuestionario = "C:/M2S2/Cuestionario.txt"
preguntas_y_respuestas = leer_cuestionario(ruta_cuestionario)

# Verificar correspondencia de preguntas
num_preguntas_cuestionario = len(preguntas_y_respuestas)
num_preguntas_excel = (ws.max_column - indice_columna_inicial + 1) // 3  # Dos columnas por pregunta

if num_preguntas_cuestionario != num_preguntas_excel:
    print("Advertencia: El número de preguntas no coincide entre el cuestionario y el Excel.")
    print("Preguntas en el cuestionario:", num_preguntas_cuestionario)
    print("Preguntas en el Excel:", num_preguntas_excel)
    # Imprimir preguntas del cuestionario
    print("\nPreguntas del cuestionario:")
    for i, preg_resp in enumerate(preguntas_y_respuestas):
        print(f"Pregunta {i + 1}: {preg_resp[0]}")  # Mostrar solo el enunciado de la pregunta

    # Imprimir preguntas del Excel
    print("\nPreguntas en el archivo Excel:")
    for col in range(indice_columna_inicial, indice_columna_inicial + num_preguntas_excel * 2, 3):
        # Asumiendo que cada pregunta tiene un encabezado en la primera fila y ocupa dos columnas
        pregunta_excel = ws.cell(row=1, column=col).value
        print(f"Pregunta (columna {openpyxl.utils.get_column_letter(col)}): {pregunta_excel}")


# Actualizar el Excel con las respuestas, correctas e incorrectas, reemplazadas por su inciso
for idx_pregunta, pregunta_respuestas in enumerate(preguntas_y_respuestas):
    # Ajustar para que la columna de respuestas avance correctamente según el número de pregunta
    columna_respuestas = indice_columna_inicial + idx_pregunta * 3  # Cada pregunta ocupa 3 columnas

    for fila in range(2, ws.max_row + 1):  # Ignorar el encabezado
        respuesta_celda = ws.cell(row=fila, column=columna_respuestas).value
        respuesta_encontrada = False

        # Buscar la respuesta del usuario en las opciones y asignar el inciso correspondiente
        for opcion in pregunta_respuestas[1:]:  # Ignorar la línea de la pregunta
            # Remover el indicador de respuesta correcta "*" para la comparación
            opcion_limpia = opcion.rsplit(" *", 1)[0].split(") ", 1)[1].strip().lower()
            if str(respuesta_celda).strip().lower() == opcion_limpia:
                inciso = opcion.split(")")[0].strip()  # Extraer el inciso
                ws.cell(row=fila, column=columna_respuestas, value=inciso)
                respuesta_encontrada = True
                break

        # Opcional: Manejar el caso en que la respuesta no coincide con ninguna opción
        if not respuesta_encontrada:
            print(f"Advertencia: La respuesta '{respuesta_celda}' en la fila {fila}, pregunta {idx_pregunta + 1} no coincide con ninguna opción proporcionada.")

# Guardar los cambios en el archivo
wb.save("C:/M2S2/Respuestas_Actualizadas.xlsx")



print("El archivo 'Respuestas_Actualizadas.xlsx' ha sido actualizado con éxito.")
